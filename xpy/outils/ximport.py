import wx,os
import platform
import csv
import xlrd
import datetime
from openpyxl import load_workbook

def GetFichierXls(nomFichier,minrow=1,maxrow=1000,mincol=1,maxcol=10):
    # pour anciennes versions de fichiers excel jusqu'à 2003
    """
    Si on veut accéder aux informations de formattage des cellules, il faut faire :
            myBook = xlrd.open_workbook('myFile.xls', formatting_info = True)
    myBook.sheets() : renvoie les feuilles du fichier sous forme de liste (objets xlrd.sheet.Sheet).
    myBook.sheet_names() : renvoie les noms des feuilles du fichier sous forme de liste.
    myBook.sheet_by_name('Feuill1') : renvoie la feuille de nom indiqué.

    Méthodes sur les objets Sheet (feuilles) :
    mySheet.name : le nom de la feuille.
    mySheet.nrows et mySheet.ncols : le nombre de lignes et de colonnes du fichier.
    mySheet.row(0) : renvoie la première ligne sous forme de liste des valeurs de type xlrd.sheet.Cell.
    mySheet.col(0) : renvoie la première colonne sous forme de liste des valeurs de type xlrd.sheet.Cell.
    mySheet.cell_type(5, 1) : le type de la valeur à la ligne 6 et la colonne 2 (origine à 0).

    mySheet.cell(5, 1) : la cellule à la ligne 6 et la colonne 2 (origine à 0), objet de type xlrd.sheet.Cell.

    Méthodes sur les objets xlrd.sheet.Cell :
    myCell.value : renvoie la valeur.
    myCell.xf_index : renvoie l'index de formattage, voir ci-dessous.
    """
    wk = xlrd.open_workbook(nomFichier)
    ws = wk.sheet_by_index(0) # renvoie la première feuille.
    lstDonnees = []
    for row in range(minrow,maxrow):
        ligne = []
        sansNull = []
        for col in range(mincol,maxcol):
            try:
                value = ws.cell_value(row - 1, col - 1)
                cellType = ws.cell_type(row-1, col-1)
                if cellType == xlrd.sheet.XL_CELL_DATE:
                    value = datetime.datetime(*xlrd.xldate.xldate_as_tuple(value, wk.datemode))
            except: value = None
            ligne.append(value)
            if value: sansNull.append(value)
        if len(sansNull)>0:
            lstDonnees.append(ligne)
    return lstDonnees

def GetFichierXlsx(nomFichier,minrow=1,maxrow=1000,mincol=1,maxcol=10):
    #get handle on existing file
    wk = load_workbook(filename=nomFichier)
    #get active worksheet or wk['some_worksheet']
    ws = wk.active
    #loop through range values
    lstDonnees = []
    for values in ws.iter_rows(min_row=minrow,max_row=maxrow,min_col=mincol,max_col=maxcol,values_only=True):
        sansNull = [x for x in values if x]
        if len(sansNull)>0:
            lstDonnees.append(values)
    return lstDonnees

def GetFichierCsv(nomFichier,delimiter="\t",detect=True):
    if platform.system() == "Windows":
        nomFichier = nomFichier.replace("/", "\\")
    # ouverture du fichier en lecture seule
    try:
        fichier = open(nomFichier, "rt")
    except Exception as err :
        wx.MessageBox("Erreur d'accès au fichier\n\nAppel: %s\nErreur: %d, %s"%(nomFichier,err.args[0],err.args[1]))
        return []
    # csv.reader est la fonction qui lit le fichier ouvert
    donnees = [x for x in csv.reader(fichier,delimiter=delimiter)]
    fichier.close()
    if detect and len(donnees[0]) == 1:
        # le séparateur n'etait pas le bon ou inexistant, essais avec ';'
        donnees = GetFichierCsv(nomFichier, delimiter=";",detect=False)
    return donnees

if __name__ == '__main__':
    app = wx.App(0)
    os.chdir("..")
    donnees = GetFichierCsv("../srcNoestock/Versions.txt")
    #donnees = GetFichierXls("c:/temp/FichierTest.xls")

    print(donnees[0])
    app.MainLoop()

